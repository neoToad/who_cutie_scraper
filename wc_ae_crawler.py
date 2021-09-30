import requests
from bs4 import BeautifulSoup
from selenium.webdriver import Chrome
from lxml import html
import os
import time
import re
from openpyxl import Workbook, load_workbook

from dotenv import load_dotenv


wb = load_workbook('available_inventory.xlsx')
ws1 = wb.active

load_dotenv()

driver = Chrome(executable_path='chromedriver.exe')
driver.get('https://whocutie.aliexpress.com/store/all-wholesale-products/2996039.html?spm=a2g0o.store_pc_home" \
      ".pcShopHead_6241067.99')

driver.maximize_window()

driver.find_element_by_id('fm-login-id').send_keys(os.getenv('EMAIL'))

driver.find_element_by_id('fm-login-password').send_keys(os.getenv('PASSWORD'))

driver.find_element_by_class_name('fm-button').click()

time.sleep(5)
html = driver.page_source

soup = BeautifulSoup(html, "lxml")

page_items = soup.find("ul", {"class": "items-list"})
items = page_items.find_all("li", {"class": "item"})

item_number = 0

for item in items:
    item_number += 1
    detail = item.find('div', {"class": "detail"})

    # Link Detail
    link = detail.find('a')
    name = link['title']
    item_page = link['href'][2:]
    print('\n')
    print(name)
    print(item_page)

    # ws1.cell(row=item_number + 1, column=1).value = item_number
    # ws1.cell(row=item_number + 1, column=2).value = name
    # ws1.cell(row=item_number + 1, column=3).value = item_page

    # price
    cost = detail.find('div', {"class": "cost"})
    price = re.findall(r"([\d.]*\d+)", cost.text)
    print(price)

    num_of_hist_orders = detail.find('div', {"class": "recent-order"})
    print(num_of_hist_orders.text)

    # ws1.cell(row=item_number + 1, column=4).value = price[0]
    # ws1.cell(row=item_number + 1, column=5).value = num_of_hist_orders.text

    # images
    image_div = item.find('div', {"class": "img"})
    image = image_div.find('img')

    print(image['image-src'])
    print(image['image-src'][:-12])
    print(image['alt'])

    # ws1.cell(row=item_number + 1, column=6).value = image['image-src']
    # ws1.cell(row=item_number + 1, column=7).value = image['image-src'][:-12]
    # ws1.cell(row=item_number + 1, column=8).value = image['alt']


driver.quit()
wb.save(filename='available_inventory.xlsx')